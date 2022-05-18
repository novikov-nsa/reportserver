import configparser
import datetime
from lxml import etree
import openpyxl as xlsx
import postgres as pg
from jinja2 import Template
import logging
import treport.logger as lg

XSD_FILE_NAME = 'treports.xsd'

def get_config(path_to_inifile):
    config = configparser.ConfigParser()
    config.read(path_to_inifile)
    login = config.get('database', 'login')
    password = config.get('database', 'password')
    host = config.get('database', 'host')
    port = config.get('database', 'port')
    database = config.get('database', 'database')
    db_url = f'postgresql://{login}:{password}@{host}:{port}/{database}'
    path_to_params_reports_file = config.get('report', 'params_reports')
    return db_url, path_to_params_reports_file


def get_errors_validation(xml_doc, xsd_doc):
    xmlschema = etree.XMLSchema(xsd_doc)
    log = xmlschema.error_log
    #return xmlschema.assertValid(xml_doc)
    return xmlschema.assert_(xml_doc)



class Report():
    codeReport = ''
    nameReport = ''
    template = ''
    outDir = ''
    params_report = {}
    xml_validation_result : bool = None
    report_file_name = ''
    reportPages = {}
    '''reportPages - атрибут, в котором хранятся сведения о листе отчета. Атрибут является словарем и имеет следующие ключи:
    pageName - имя листа в xlsx-файле
    sqlFile - путь к файлу, в котором хранится SQL-запрос
    headerRows - количество строк на листе, которые выделены под заголовок
    ignoredColumns - список номеров колонок, которые необходимо игнорировать при формировании отчета
    sqlText - текст SQL-запроса, который сформирован на основании текста в файле с подставленными значениями параметров запроса
    sqlResult - результат выполнения SQL-запроса
    '''
    contentReport = None
    '''
    Контент сформированного отчета. 
    '''
    paramValues = {}
    dbUrl = ''
    logger = logging.Logger


    def __init__(self, report_code, path_to_params_reports_file, param_values, db_url):
        self.logger = lg.get_logger(__name__)
        self.codeReport = report_code
        self.get_params_report(path_to_params_reports_file)
        if self.xml_validation_result:
            self.paramValues = param_values
            self.generate_sql()
            self.dbUrl = db_url
            self.get_dataset()
            self.generate_report()
        else:
            self.logger.info(f'Проверка XML-файла {path_to_params_reports_file} проведена, файл имеет ошибки')


    def validate_params_report(self, xml_doc, xsd_doc):
        xmlschema = etree.XMLSchema(xsd_doc)
        return xmlschema.validate(xml_doc)

    def generate_file_name(self, filename_rule):
        result_file_name = 'Итоговый_файл_' + str(datetime.datetime.now()) + '.xlsx'
        self.logger.info('Имя файла '+result_file_name+' сформировано')
        return result_file_name

    def get_params_report(self, path_to_params_reports_file):
        xsd_f = open(XSD_FILE_NAME)
        xml_f = open(path_to_params_reports_file)
        xsd_doc = etree.parse(xsd_f)
        xml_doc = etree.parse(xml_f)
        self.xml_validation_result = self.validate_params_report(xml_doc, xsd_doc)
        if self.xml_validation_result:
            self.logger.info(f'Проверка XML-файла {path_to_params_reports_file} проведена, файл корректен')
            xml_root = xml_doc.getroot()[0]

            for item_report in xml_root:
                if item_report.attrib['codeReport'] == self.codeReport:
                    self.logger.info(f'Чтение блока параметров отчета {self.codeReport}')

                    self.nameReport = item_report[0].text       # Наименование отчета
                    self.logger.info(f'Отчет {self.nameReport}')

                    self.template = item_report[1].text         # Путь к файлу-шаблону отчета
                    self.logger.info(f'Шаблон {self.template}')

                    self.outDir = item_report[2].text           # Путь к каталогу, где будет сохраняться отчет
                    self.logger.info(f'Путь к каталогу, где будет сохраняться отчет: {self.template}')

                    file_name_rule = item_report[3]             # Правило конструирования имени файла
                    self.report_file_name = self.generate_file_name(file_name_rule)
                    self.logger.info(f'Имя файла отчета: {self.template}')


                    params = item_report[4]                     # Параметры отчета
                    self.logger.info(f'Параметры отчета {self.codeReport}')
                    for item_params in params:
                        parametr_code = item_params.attrib['id']
                        self.params_report[parametr_code] = {}
                        for parametr in item_params:
                            self.params_report[parametr_code][parametr.tag] = parametr.text
                            self.logger.info(f'Параметр: {parametr_code}, значение: {parametr.text}')

                    report_pages = item_report[5]               # Свойства каждой страницы отчета
                    self.logger.info(f'Свойства страниц отчета {self.codeReport}')

                    for item_page in report_pages:
                        page_code = item_page.attrib['codePage']
                        self.reportPages[page_code] = {}
                        for parametr_page in item_page:
                            if parametr_page.tag == 'ignoredColumns':
                                igonred_columns = parametr_page.text.split(',')
                                int_ignored_columns = []
                                for item in igonred_columns:
                                    int_ignored_columns.append(int(item))
                                self.reportPages[page_code][parametr_page.tag] = int_ignored_columns

                            if parametr_page.tag == 'headerRows':
                                self.reportPages[page_code][parametr_page.tag] = int(parametr_page.text)

                            if parametr_page.tag not in ('ignoredColumns', 'headerRows'):
                                self.reportPages[page_code][parametr_page.tag] = parametr_page.text
                        self.logger.info(f'Страница {page_code} параметр {parametr_page} значение {parametr_page.text}')



    def generate_sql(self):
        for report_page in self.reportPages:
            f = open(self.reportPages[report_page]['sqlFile'])
            template = Template(f.read())
            self.reportPages[report_page]['sqlText'] = template.render(self.paramValues)
            self.logger.info(f'SQL-файл отчета {self.codeReport} для страницы {report_page} сформирован')



    def get_dataset(self):
        db = pg.Postgres(url=self.dbUrl)
        for report_page in self.reportPages:
            sql_text = self.reportPages[report_page]['sqlText']
            data_set_page = db.all(sql_text)
            self.reportPages[report_page]['sqlResult'] = data_set_page
            self.logger.info(f'Запрос к БД для формирвоания отчета {self.codeReport} для страницы {report_page} выполнен')

    def delete_rows(self, ws, max_data_rows):
        max_sheet_rows = ws.max_row
        rows_to_delete = max_sheet_rows - max_data_rows
        ws.delete_rows(max_data_rows + 1, rows_to_delete)

    def generate_table(self, ws, header, list_page, code_page):
        '''
        Функция формирования файла в формате OpenXML (xlsx) по результатам SQL-запроса

        :param ws: лист рабочей книги эксель
        :param header: количество строк в таблице, отведенных под заголовок
        :param list_page: результат запроса к БД
        :param name_sheet: наименование листа рабочей книги эксель
        :return:
        '''
        counter = 0
        for item in list_page:
            counter += 1
            line_number = header + counter
            col_count = 0
            for field in item:
                col_count += 1
                if col_count not in self.reportPages[code_page]['ignoredColumns']:
                    ws.cell(row=line_number, column=col_count).value = item[col_count - 1]

        self.logger.info(f'Лист {code_page} отчета {self.codeReport} заполнен')

        self.delete_rows(ws, counter + header)

    def generate_report(self):
        wb = xlsx.load_workbook(filename=self.template, read_only=False, keep_vba=False, data_only=False)
        self.logger.info(f'Формирование файла отчета {self.codeReport} в формате MS Excel начато')

        for report_page in self.reportPages:
            ws = wb[self.reportPages[report_page]['pageName']]
            header = self.reportPages[report_page]['headerRows']
            dataset_page = self.reportPages[report_page]['sqlResult']
            self.generate_table(ws, header, dataset_page, report_page)
        self.contentReport = wb




def main(report_code, parameters):
    db_url, path_to_params_reports_file = get_config('/Users/sergejnovikov/PycharmProjects/reportserver/treport/main.ini')
    report = Report(report_code, path_to_params_reports_file, parameters, db_url)

    if report.xml_validation_result:
        report.contentReport.save(report.outDir+report.report_file_name)
        report.logger.info(f'Файл {report.outDir+report.report_file_name} сохранен')
    else:
        print('Файл не корректен')

